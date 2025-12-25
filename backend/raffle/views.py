from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
from django.conf import settings
from django.template.loader import render_to_string
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import RaffleEvent, Prize, RaffleParticipant, RaffleLog
from .serializers import (
    RaffleEventSerializer, PrizeSerializer, RaffleParticipantSerializer, RaffleLogSerializer
)
from .algorithms import RaffleSelector
from core.permissions import IsOrgAdminOrReadOnly, IsStaffOrReadOnly, IsRaffleOperator
from core.utils import create_audit_log, ImportProcessor
from core.models import Organization, Event, Department
from teams.models import Participant
from rest_framework.permissions import IsAuthenticated
from config.pagination import StandardResultsSetPagination
import pandas as pd
import io


class RaffleEventViewSet(viewsets.ModelViewSet):
    serializer_class = RaffleEventSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['org', 'event']
    search_fields = ['name', 'description']
    ordering_fields = ['created_at']
    ordering = ['-created_at']
    permission_classes = [IsAuthenticated, IsOrgAdminOrReadOnly]
    
    def dispatch(self, request, *args, **kwargs):
        # #region agent log
        import json
        try:
            with open('/Users/jr/HOS-Develop/HIT-Project/HosProject/.cursor/debug.log', 'a') as f:
                f.write(json.dumps({'location': 'raffle/views.py:dispatch', 'message': 'RaffleEventViewSet dispatch called', 'data': {'method': request.method, 'path': request.path, 'full_path': request.get_full_path(), 'view_action': kwargs.get('action'), 'pk': kwargs.get('pk')}, 'timestamp': int(__import__('time').time() * 1000), 'sessionId': 'debug-session', 'runId': 'run1', 'hypothesisId': 'D'}) + '\n')
        except: pass
        # #endregion
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        org_id = getattr(self.request, 'org_id', None)
        if org_id:
            queryset = RaffleEvent.objects.select_related('org', 'event').filter(org_id=org_id)
            event_id = self.request.query_params.get('event')
            if event_id:
                queryset = queryset.filter(event_id=event_id)
            return queryset
        if self.request.user.is_superadmin():
            return RaffleEvent.objects.select_related('org', 'event').all()
        return RaffleEvent.objects.select_related('org', 'event').none()
    
    def create(self, request, *args, **kwargs):
        """Override create to ensure proper serialization"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        
        # Refresh the instance to ensure related objects are loaded
        instance = serializer.instance
        instance.refresh_from_db()
        
        # Re-serialize with fresh data including related objects
        output_serializer = self.get_serializer(instance)
        headers = self.get_success_headers(output_serializer.data)
        return Response(output_serializer.data, status=status.HTTP_201_CREATED, headers=headers)
    
    def perform_create(self, serializer):
        """Override to add audit log on create"""
        raffle_event = serializer.save()
        
        # Refresh from DB to ensure related objects are loaded
        raffle_event.refresh_from_db()
        
        org = raffle_event.org
        
        # Fallback to get org from request context if not set
        if not org:
            org_id = getattr(self.request, 'org_id', None)
            if org_id:
                try:
                    org = Organization.objects.get(id=org_id)
                except Organization.DoesNotExist:
                    org = None
        
        # Prepare changes for audit log
        changes = {'name': raffle_event.name}
        if raffle_event.event_id:
            changes['event_id'] = raffle_event.event_id
        
        try:
            create_audit_log(
                user=self.request.user,
                org=org,
                action='create',
                model='RaffleEvent',
                object_id=raffle_event.id,
                changes=changes,
                request=self.request
            )
        except Exception as e:
            # Log error but don't fail the creation
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Failed to create audit log for RaffleEvent {raffle_event.id}: {e}')
    
    def perform_update(self, serializer):
        """Override to add audit log on update"""
        old_raffle = self.get_object()
        raffle_event = serializer.save()
        org = raffle_event.org
        
        # Track changes
        changes = {}
        if old_raffle.name != raffle_event.name:
            changes['name'] = {'old': old_raffle.name, 'new': raffle_event.name}
        if old_raffle.description != raffle_event.description:
            changes['description'] = {'old': old_raffle.description, 'new': raffle_event.description}
        if old_raffle.event_id != raffle_event.event_id:
            changes['event_id'] = {'old': old_raffle.event_id, 'new': raffle_event.event_id}
        
        create_audit_log(
            user=self.request.user,
            org=org,
            action='update',
            model='RaffleEvent',
            object_id=raffle_event.id,
            changes=changes,
            request=self.request
        )
    
    def destroy(self, request, *args, **kwargs):
        """Override to require reason and add audit log"""
        reason = request.data.get('reason', '')
        if not reason or len(reason.strip()) < 10:
            return Response(
                {'error': 'Reason is required and must be at least 10 characters'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        raffle_event = self.get_object()
        org = raffle_event.org
        event_name = raffle_event.name
        
        # Create audit log before deletion
        create_audit_log(
            user=request.user,
            org=org,
            action='delete',
            model='RaffleEvent',
            object_id=raffle_event.id,
            changes={'name': event_name, 'reason': reason},
            request=request
        )
        
        # Delete the object
        self.perform_destroy(raffle_event)
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=True, methods=['get'], url_path='list-eligible-participants')
    def list_eligible_participants(self, request, pk=None):
        """
        List all eligible participants for this raffle event.
        Returns all participants in the event who have is_raffle_eligible=True.
        """
        # #region agent log
        import json
        try:
            with open('/Users/jr/HOS-Develop/HIT-Project/HosProject/.cursor/debug.log', 'a') as f:
                f.write(json.dumps({'location': 'raffle/views.py:155', 'message': 'list_eligible_participants action called', 'data': {'pk': pk, 'method': request.method, 'path': request.path, 'full_path': request.get_full_path()}, 'timestamp': int(__import__('time').time() * 1000), 'sessionId': 'debug-session', 'runId': 'run1', 'hypothesisId': 'C'}) + '\n')
        except: pass
        # #endregion
        raffle_event = self.get_object()
        # #region agent log
        try:
            with open('/Users/jr/HOS-Develop/HIT-Project/HosProject/.cursor/debug.log', 'a') as f:
                f.write(json.dumps({'location': 'raffle/views.py:162', 'message': 'RaffleEvent retrieved', 'data': {'raffle_event_id': raffle_event.id if raffle_event else None, 'event_id': raffle_event.event.id if raffle_event and raffle_event.event else None}, 'timestamp': int(__import__('time').time() * 1000), 'sessionId': 'debug-session', 'runId': 'run1', 'hypothesisId': 'C'}) + '\n')
        except: pass
        # #endregion
        event = raffle_event.event
        
        # Get all participants in the event who are eligible
        participants = Participant.objects.filter(
            event_id=event.id,
            is_raffle_eligible=True
        ).select_related('department', 'org').prefetch_related('team_memberships__team')
        
        # Serialize participants
        from teams.serializers import ParticipantSerializer
        serializer = ParticipantSerializer(participants, many=True)
        
        return Response({
            'success': True,
            'count': participants.count(),
            'results': serializer.data
        })
    
    @action(detail=True, methods=['post'], url_path='reset-all-prizes')
    def reset_all_prizes(self, request, pk=None):
        """Reset all prizes in this raffle event"""
        raffle_event = self.get_object()
        reason = request.data.get('reason', '').strip()
        
        if not reason or len(reason) < 10:
            return Response(
                {'error': 'Reason is required and must be at least 10 characters'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get all prizes for this raffle event
        prizes = raffle_event.prizes.all()
        total_deleted = 0
        prize_details = []
        all_participant_ids = set()
        
        # Delete all RaffleParticipant records for all prizes and collect participant IDs
        for prize in prizes:
            deleted_count = prize.selected_participants.count()
            if deleted_count > 0:
                participant_ids = list(prize.selected_participants.values_list('participant_id', flat=True))
                all_participant_ids.update(participant_ids)
                prize.selected_participants.all().delete()
                total_deleted += deleted_count
                prize_details.append({
                    'prize_id': prize.id,
                    'prize_name': prize.name,
                    'deleted_count': deleted_count
                })
        
        # Restore raffle eligibility for all participants who won any prize in this raffle event
        if all_participant_ids:
            Participant.objects.filter(id__in=all_participant_ids).update(is_raffle_eligible=True)
        
        # Create audit log
        create_audit_log(
            user=request.user,
            org=raffle_event.org,
            action='delete',
            model='RaffleParticipant',
            changes={
                'raffle_event_id': raffle_event.id,
                'raffle_event_name': raffle_event.name,
                'total_deleted_count': total_deleted,
                'prizes': prize_details,
                'reason': reason,
                'reset_all': True
            },
            request=request
        )
        
        return Response({
            'success': True,
            'message': f'Reset {total_deleted} winner(s) across {len(prize_details)} prize(s)',
            'total_deleted_count': total_deleted,
            'prizes_affected': len(prize_details)
        })
    
    @action(detail=True, methods=['get'], url_path='public-info', permission_classes=[])
    def public_info(self, request, pk=None):
        """Get raffle event info (public, no auth required)"""
        try:
            raffle_event = RaffleEvent.objects.select_related('org', 'event').get(id=pk)
        except RaffleEvent.DoesNotExist:
            return Response(
                {'error': 'Raffle event not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        serializer = self.get_serializer(raffle_event)
        return Response(serializer.data)


class PrizeViewSet(viewsets.ModelViewSet):
    serializer_class = PrizeSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['raffle_event', 'round_number']
    search_fields = ['name']
    ordering_fields = ['round_number', 'created_at']
    ordering = ['round_number', 'name']
    permission_classes = [IsAuthenticated, IsOrgAdminOrReadOnly]
    
    def get_queryset(self):
        org_id = getattr(self.request, 'org_id', None)
        if org_id:
            return Prize.objects.select_related('raffle_event', 'raffle_event__org').filter(raffle_event__org_id=org_id)
        if self.request.user.is_superadmin():
            return Prize.objects.select_related('raffle_event', 'raffle_event__org').all()
        return Prize.objects.select_related('raffle_event', 'raffle_event__org').none()
    
    def perform_create(self, serializer):
        """Override to add audit log on create"""
        prize = serializer.save()
        prize.refresh_from_db()
        org = prize.raffle_event.org if prize.raffle_event else None
        
        create_audit_log(
            user=self.request.user,
            org=org,
            action='create',
            model='Prize',
            object_id=prize.id,
            changes={'name': prize.name, 'round_number': prize.round_number, 'raffle_event_id': prize.raffle_event_id},
            request=self.request
        )
    
    def perform_update(self, serializer):
        """Override to add audit log on update"""
        old_prize = self.get_object()
        prize = serializer.save()
        prize.refresh_from_db()
        org = prize.raffle_event.org if prize.raffle_event else None
        
        # Track changes
        changes = {}
        if old_prize.name != prize.name:
            changes['name'] = {'old': old_prize.name, 'new': prize.name}
        if old_prize.round_number != prize.round_number:
            changes['round_number'] = {'old': old_prize.round_number, 'new': prize.round_number}
        if old_prize.quantity != prize.quantity:
            changes['quantity'] = {'old': old_prize.quantity, 'new': prize.quantity}
        
        create_audit_log(
            user=self.request.user,
            org=org,
            action='update',
            model='Prize',
            object_id=prize.id,
            changes=changes,
            request=self.request
        )
    
    def destroy(self, request, *args, **kwargs):
        """Override to require reason and add audit log"""
        reason = request.data.get('reason', '')
        if not reason or len(reason.strip()) < 10:
            return Response(
                {'error': 'Reason is required and must be at least 10 characters'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        prize = self.get_object()
        org = prize.raffle_event.org if prize.raffle_event else None
        prize_name = prize.name
        
        # Create audit log before deletion
        create_audit_log(
            user=request.user,
            org=org,
            action='delete',
            model='Prize',
            object_id=prize.id,
            changes={'name': prize_name, 'reason': reason},
            request=request
        )
        
        # Delete the object
        self.perform_destroy(prize)
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=True, methods=['post'], url_path='reset')
    def reset_prize(self, request, pk=None):
        """Reset winners for a single prize"""
        prize = self.get_object()
        reason = request.data.get('reason', '').strip()
        
        if not reason or len(reason) < 10:
            return Response(
                {'error': 'Reason is required and must be at least 10 characters'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get count and participant IDs before deletion for audit log
        deleted_count = prize.selected_participants.count()
        participant_ids = list(prize.selected_participants.values_list('participant_id', flat=True))
        
        # Delete all RaffleParticipant records for this prize
        prize.selected_participants.all().delete()
        
        # Restore raffle eligibility for all participants who won this prize
        if participant_ids:
            Participant.objects.filter(id__in=participant_ids).update(is_raffle_eligible=True)
        
        # Create audit log
        create_audit_log(
            user=request.user,
            org=prize.raffle_event.org if prize.raffle_event else None,
            action='delete',
            model='RaffleParticipant',
            changes={
                'prize_id': prize.id,
                'prize_name': prize.name,
                'deleted_count': deleted_count,
                'reason': reason,
                'reset': True
            },
            request=request
        )
        
        return Response({
            'success': True,
            'message': f'Reset {deleted_count} winner(s) for prize "{prize.name}"',
            'deleted_count': deleted_count
        })
    
    @action(detail=True, methods=['post'], url_path='select')
    def select_winners(self, request, pk=None):
        """Select winners for this prize"""
        prize = self.get_object()
        quantity = request.data.get('quantity', prize.quantity)
        rules = request.data.get('rules', prize.rules)
        
        # Merge with prize rules and add no_repeat_prize from raffle_event
        merged_rules = {**prize.rules, **rules}
        merged_rules['no_repeat_prize'] = prize.raffle_event.no_repeat_prize
        
        # Create selector
        selector = RaffleSelector(prize, merged_rules)
        
        # Select winners
        result = selector.select(quantity=quantity)
        
        if not result['success']:
            return Response(result, status=status.HTTP_400_BAD_REQUEST)
        
        # Create RaffleParticipant records and disable eligibility
        created_count = 0
        participants_to_disable = []
        for winner in result['winners']:
            raffle_participant, created = RaffleParticipant.objects.get_or_create(
                prize=prize,
                participant=winner,
                defaults={'seed_value': result['seed']}
            )
            if created:
                created_count += 1
                # Mark participant for eligibility disable
                participants_to_disable.append(winner)
        
        # Disable raffle eligibility for all participants who won
        if participants_to_disable:
            Participant.objects.filter(
                id__in=[p.id for p in participants_to_disable]
            ).update(is_raffle_eligible=False)
        
        # Create log
        RaffleLog.objects.create(
            raffle_event=prize.raffle_event,
            prize=prize,
            seed=result['seed'],
            rule_snapshot=result['rule_snapshot'],
            result={
                **result['result'],
                'selected_participants': [p.id for p in result['winners']]
            }
        )
        
        # Audit log
        create_audit_log(
            user=request.user,
            org=prize.raffle_event.org,
            action='create',
            model='RaffleParticipant',
            changes={
                'prize_id': prize.id,
                'selected_count': created_count,
                'seed': result['seed']
            },
            request=request
        )
        
        # Broadcast via WebSocket (will be implemented)
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f'raffle_{prize.raffle_event.id}',
                {
                    'type': 'raffle_result',
                    'prize_id': prize.id,
                    'winners': [
                        {
                            'id': p.id,
                            'name': p.name,
                            'department': p.department.name if p.department else None
                        }
                        for p in result['winners']
                    ],
                    'seed': result['seed']
                }
            )
        
        return Response({
            'success': True,
            'winners': [
                {
                    'id': p.id,
                    'name': p.name,
                    'department': p.department.name if p.department else None
                }
                for p in result['winners']
            ],
            'seed': result['seed'],
            'created_count': created_count
        })
    
    @action(detail=True, methods=['post'], url_path='add-participants')
    def add_participants(self, request, pk=None):
        """Manually add participants to this prize"""
        prize = self.get_object()
        participant_ids = request.data.get('participant_ids', [])
        
        if not participant_ids or not isinstance(participant_ids, list):
            return Response(
                {'error': 'participant_ids is required and must be a list'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate participants exist and belong to the same event
        from teams.models import Participant
        try:
            participants = Participant.objects.filter(
                id__in=participant_ids,
                event_id=prize.raffle_event.event_id,
                org_id=prize.raffle_event.org_id
            )
            if participants.count() != len(participant_ids):
                return Response(
                    {'error': 'Some participants not found or do not belong to this event'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Exception as e:
            return Response(
                {'error': f'Error validating participants: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create RaffleParticipant records and disable eligibility
        created_count = 0
        participants_to_disable = []
        import hashlib
        from datetime import datetime
        seed = hashlib.sha256(f"{datetime.now().isoformat()}_{prize.id}_manual".encode()).hexdigest()
        
        for participant in participants:
            raffle_participant, created = RaffleParticipant.objects.get_or_create(
                prize=prize,
                participant=participant,
                defaults={'seed_value': seed}
            )
            if created:
                created_count += 1
                # Mark participant for eligibility disable
                participants_to_disable.append(participant)
        
        # Disable raffle eligibility for all participants who won
        if participants_to_disable:
            Participant.objects.filter(
                id__in=[p.id for p in participants_to_disable]
            ).update(is_raffle_eligible=False)
        
        # Audit log
        create_audit_log(
            user=request.user,
            org=prize.raffle_event.org,
            action='create',
            model='RaffleParticipant',
            changes={
                'prize_id': prize.id,
                'selected_count': created_count,
                'participant_ids': participant_ids,
                'manual': True
            },
            request=request
        )
        
        return Response({
            'success': True,
            'created_count': created_count,
            'participants': [
                {
                    'id': p.id,
                    'name': p.name,
                    'department': p.department.name if p.department else None
                }
                for p in participants
            ]
        })
    
    @action(detail=False, methods=['get'], url_path='public-list', permission_classes=[])
    def public_list(self, request):
        """List prizes for a raffle_event (public, no auth required)"""
        raffle_event_id = request.query_params.get('raffle_event')
        if not raffle_event_id:
            return Response(
                {'error': 'raffle_event parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            raffle_event = RaffleEvent.objects.get(id=raffle_event_id)
        except RaffleEvent.DoesNotExist:
            return Response(
                {'error': 'Raffle event not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        prizes = Prize.objects.filter(raffle_event_id=raffle_event_id).select_related('raffle_event')
        serializer = self.get_serializer(prizes, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'], url_path='public-detail', permission_classes=[])
    def public_detail(self, request, pk=None):
        """Get prize detail (public, no auth required)"""
        try:
            prize = Prize.objects.select_related('raffle_event').get(id=pk)
        except Prize.DoesNotExist:
            return Response(
                {'error': 'Prize not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        serializer = self.get_serializer(prize)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'], url_path='public-select', permission_classes=[IsAuthenticated, IsRaffleOperator])
    def public_select_winners(self, request, pk=None):
        """Select winners for this prize (requires authentication and raffle operator role) - DOES NOT SAVE"""
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"public_select_winners called for prize {pk} - NOT SAVING")
        
        try:
            prize = Prize.objects.select_related('raffle_event', 'raffle_event__event').get(id=pk)
        except Prize.DoesNotExist:
            return Response(
                {'error': 'Prize not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        quantity = request.data.get('quantity', prize.quantity)
        rules = request.data.get('rules', prize.rules)
        
        # Merge with prize rules and add no_repeat_prize from raffle_event
        merged_rules = {**prize.rules, **rules}
        merged_rules['no_repeat_prize'] = prize.raffle_event.no_repeat_prize
        
        # Create selector
        selector = RaffleSelector(prize, merged_rules)
        
        # Select winners (but don't save yet)
        result = selector.select(quantity=quantity)
        
        if not result['success']:
            return Response(result, status=status.HTTP_400_BAD_REQUEST)
        
        # IMPORTANT: Do NOT create RaffleParticipant records here
        # This endpoint only selects winners, does not save them
        logger.info(f"Selected {len(result['winners'])} winners for prize {pk} - NOT SAVED")
        
        # Return winners data without saving
        return Response({
            'success': True,
            'winners': [
                {
                    'id': p.id,
                    'name': p.name,
                    'department': p.department.name if p.department else None
                }
                for p in result['winners']
            ],
            'seed': result['seed'],
            'rule_snapshot': result['rule_snapshot'],
            'result': result['result']
        })
    
    @action(detail=True, methods=['post'], url_path='public-save-winners', permission_classes=[IsAuthenticated, IsRaffleOperator])
    def public_save_winners(self, request, pk=None):
        """Save selected winners for this prize (requires authentication and raffle operator role)"""
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"public_save_winners called for prize {pk} - SAVING NOW")
        
        try:
            prize = Prize.objects.select_related('raffle_event', 'raffle_event__event').get(id=pk)
        except Prize.DoesNotExist:
            return Response(
                {'error': 'Prize not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get winners data from request
        participant_ids = request.data.get('participant_ids', [])
        seed = request.data.get('seed', '')
        rule_snapshot = request.data.get('rule_snapshot', {})
        result_data = request.data.get('result', {})
        
        if not participant_ids:
            return Response(
                {'error': 'participant_ids is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        logger.info(f"Saving {len(participant_ids)} winners for prize {pk}")
        
        # Validate participants
        try:
            participants = Participant.objects.filter(
                id__in=participant_ids,
                event_id=prize.raffle_event.event_id,
                org_id=prize.raffle_event.org_id
            )
            if participants.count() != len(participant_ids):
                return Response(
                    {'error': 'Some participants not found or do not belong to this event'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Exception as e:
            return Response(
                {'error': f'Error validating participants: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create RaffleParticipant records and disable eligibility
        created_count = 0
        participants_to_disable = []
        for participant in participants:
            raffle_participant, created = RaffleParticipant.objects.get_or_create(
                prize=prize,
                participant=participant,
                defaults={'seed_value': seed}
            )
            if created:
                created_count += 1
                # Mark participant for eligibility disable
                participants_to_disable.append(participant)
        
        # Disable raffle eligibility for all participants who won
        if participants_to_disable:
            Participant.objects.filter(
                id__in=[p.id for p in participants_to_disable]
            ).update(is_raffle_eligible=False)
        
        # Create log (without user for public endpoint)
        RaffleLog.objects.create(
            raffle_event=prize.raffle_event,
            prize=prize,
            seed=seed,
            rule_snapshot=rule_snapshot,
            result={
                **result_data,
                'selected_participants': participant_ids
            }
        )
        
        # Broadcast via WebSocket
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        from datetime import datetime
        
        channel_layer = get_channel_layer()
        if channel_layer:
            # Broadcast raffle_result
            async_to_sync(channel_layer.group_send)(
                f'raffle_{prize.raffle_event.id}',
                {
                    'type': 'raffle_result',
                    'prize_id': prize.id,
                    'winners': [
                        {
                            'id': p.id,
                            'name': p.name,
                            'department': p.department.name if p.department else None
                        }
                        for p in participants
                    ],
                    'seed': seed
                }
            )
            
            # Broadcast winners_update for WinnersList
            async_to_sync(channel_layer.group_send)(
                f'raffle_{prize.raffle_event.id}',
                {
                    'type': 'winners_update',
                    'raffle_event_id': prize.raffle_event.id,
                    'winners': [
                        {
                            'id': p.id,
                            'participant_name': p.name,
                            'participant': p.hospital_id,
                            'selected_at': datetime.now().isoformat()
                        }
                        for p in participants
                    ],
                    'timestamp': datetime.now().isoformat()
                }
            )
        
        return Response({
            'success': True,
            'created_count': created_count,
            'participants': [
                {
                    'id': p.id,
                    'name': p.name,
                    'department': p.department.name if p.department else None
                }
                for p in participants
            ]
        })


class RaffleParticipantViewSet(viewsets.ModelViewSet):
    serializer_class = RaffleParticipantSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['prize', 'participant', 'prize__raffle_event']
    search_fields = ['participant__name', 'prize__name']
    ordering_fields = ['selected_at']
    ordering = ['-selected_at']
    permission_classes = [IsAuthenticated, IsOrgAdminOrReadOnly]
    pagination_class = StandardResultsSetPagination
    
    def get_queryset(self):
        org_id = getattr(self.request, 'org_id', None)
        queryset = RaffleParticipant.objects.select_related('prize', 'prize__raffle_event', 'prize__raffle_event__org', 'participant')
        if org_id:
            queryset = queryset.filter(prize__raffle_event__org_id=org_id)
        elif self.request.user.is_superadmin():
            pass  # Return all
        else:
            return RaffleParticipant.objects.none()
        
        # Filter by raffle_event if provided
        raffle_event_id = self.request.query_params.get('raffle_event')
        if raffle_event_id:
            queryset = queryset.filter(prize__raffle_event_id=raffle_event_id)
        
        return queryset
    
    def perform_create(self, serializer):
        """Override to add audit log on create"""
        raffle_participant = serializer.save()
        raffle_participant.refresh_from_db()
        org = raffle_participant.prize.raffle_event.org if raffle_participant.prize and raffle_participant.prize.raffle_event else None
        
        create_audit_log(
            user=self.request.user,
            org=org,
            action='create',
            model='RaffleParticipant',
            object_id=raffle_participant.id,
            changes={
                'prize_id': raffle_participant.prize_id,
                'participant_id': raffle_participant.participant_id,
                'participant_name': raffle_participant.participant.name if raffle_participant.participant else None
            },
            request=self.request
        )
    
    @action(detail=False, methods=['post'], url_path='mark-printed')
    def mark_printed(self, request):
        """Mark winners as printed (bulk operation)"""
        org_id = getattr(request, 'org_id', None)
        if not org_id:
            return Response(
                {'error': 'Organization ID not found'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        winner_ids = request.data.get('winner_ids', [])
        if not winner_ids or not isinstance(winner_ids, list):
            return Response(
                {'error': 'winner_ids is required and must be a list'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get queryset filtered by org
        queryset = RaffleParticipant.objects.filter(
            id__in=winner_ids,
            prize__raffle_event__org_id=org_id
        )
        
        if queryset.count() != len(winner_ids):
            return Response(
                {'error': 'Some winners not found or do not belong to this organization'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Update print status
        from django.utils import timezone
        updated_count = queryset.update(
            is_printed=True,
            printed_at=timezone.now()
        )
        
        # Create audit log
        for winner in queryset:
            org = winner.prize.raffle_event.org if winner.prize and winner.prize.raffle_event else None
            create_audit_log(
                user=request.user,
                org=org,
                action='update',
                model='RaffleParticipant',
                object_id=winner.id,
                changes={
                    'is_printed': True,
                    'printed_at': timezone.now().isoformat()
                },
                request=request
            )
        
        return Response({
            'message': f'Marked {updated_count} winners as printed',
            'updated_count': updated_count
        })
    
    @action(detail=False, methods=['post'], url_path='unmark-printed')
    def unmark_printed(self, request):
        """Unmark winners as printed (bulk operation)"""
        org_id = getattr(request, 'org_id', None)
        if not org_id:
            return Response(
                {'error': 'Organization ID not found'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        winner_ids = request.data.get('winner_ids', [])
        if not winner_ids or not isinstance(winner_ids, list):
            return Response(
                {'error': 'winner_ids is required and must be a list'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get queryset filtered by org
        queryset = RaffleParticipant.objects.filter(
            id__in=winner_ids,
            prize__raffle_event__org_id=org_id
        )
        
        if queryset.count() != len(winner_ids):
            return Response(
                {'error': 'Some winners not found or do not belong to this organization'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Update print status
        updated_count = queryset.update(
            is_printed=False,
            printed_at=None
        )
        
        # Create audit log
        for winner in queryset:
            org = winner.prize.raffle_event.org if winner.prize and winner.prize.raffle_event else None
            create_audit_log(
                user=request.user,
                org=org,
                action='update',
                model='RaffleParticipant',
                object_id=winner.id,
                changes={
                    'is_printed': False,
                    'printed_at': None
                },
                request=request
            )
        
        return Response({
            'message': f'Unmarked {updated_count} winners as printed',
            'updated_count': updated_count
        })
    
    def perform_update(self, serializer):
        """Override to add audit log on update"""
        old_participant = self.get_object()
        raffle_participant = serializer.save()
        raffle_participant.refresh_from_db()
        org = raffle_participant.prize.raffle_event.org if raffle_participant.prize and raffle_participant.prize.raffle_event else None
        
        # Track changes
        changes = {}
        if old_participant.prize_id != raffle_participant.prize_id:
            changes['prize_id'] = {'old': old_participant.prize_id, 'new': raffle_participant.prize_id}
        if old_participant.participant_id != raffle_participant.participant_id:
            changes['participant_id'] = {'old': old_participant.participant_id, 'new': raffle_participant.participant_id}
        
        create_audit_log(
            user=self.request.user,
            org=org,
            action='update',
            model='RaffleParticipant',
            object_id=raffle_participant.id,
            changes=changes,
            request=self.request
        )
    
    def destroy(self, request, *args, **kwargs):
        """Override to require reason and add audit log"""
        reason = request.data.get('reason', '')
        if not reason or len(reason.strip()) < 10:
            return Response(
                {'error': 'Reason is required and must be at least 10 characters'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        raffle_participant = self.get_object()
        org = raffle_participant.prize.raffle_event.org if raffle_participant.prize and raffle_participant.prize.raffle_event else None
        participant_name = raffle_participant.participant.name if raffle_participant.participant else 'Unknown'
        prize_name = raffle_participant.prize.name if raffle_participant.prize else 'Unknown'
        
        # Create audit log before deletion
        create_audit_log(
            user=request.user,
            org=org,
            action='delete',
            model='RaffleParticipant',
            object_id=raffle_participant.id,
            changes={
                'participant_name': participant_name,
                'prize_name': prize_name,
                'reason': reason
            },
            request=request
        )
        
        # Delete the object
        self.perform_destroy(raffle_participant)
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=False, methods=['get'], url_path='export-pdf')
    def export_winners_pdf(self, request):
        """Export winners to PDF for printing"""
        org_id = getattr(request, 'org_id', None)
        if not org_id:
            return Response(
                {'error': 'Organization ID not found'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        raffle_event_id = request.query_params.get('raffle_event')
        if not raffle_event_id:
            return Response(
                {'error': 'Raffle Event ID is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get filter parameters
        search = request.query_params.get('search', '')
        prize_id = request.query_params.get('prize')
        only_unprinted = request.query_params.get('only_unprinted', 'false').lower() == 'true'
        start_index = request.query_params.get('start_index')
        end_index = request.query_params.get('end_index')
        
        # Get queryset (same logic as list, but without pagination)
        queryset = RaffleParticipant.objects.select_related(
            'prize', 'prize__raffle_event', 'participant', 'participant__department'
        ).filter(prize__raffle_event_id=raffle_event_id, prize__raffle_event__org_id=org_id)
        
        if search:
            queryset = queryset.filter(participant__name__icontains=search)
        
        if prize_id:
            queryset = queryset.filter(prize_id=prize_id)
        
        # Filter by print status
        if only_unprinted:
            queryset = queryset.filter(is_printed=False)
        
        # Order by selected_at descending (newest first = rank 1)
        queryset = queryset.order_by('-selected_at')
        
        # Apply range selection if provided
        if start_index or end_index:
            start = int(start_index) - 1 if start_index else 0  # Convert to 0-based index
            end = int(end_index) if end_index else None  # end_index is inclusive, so use as-is for slicing
            queryset = queryset[start:end]
        
        # Get raffle event and org info
        try:
            raffle_event = RaffleEvent.objects.select_related('event', 'org').get(id=raffle_event_id, org_id=org_id)
            event = raffle_event.event if hasattr(raffle_event, 'event') else None
            org = raffle_event.org if hasattr(raffle_event, 'org') else None
            raffle_name = raffle_event.name
        except RaffleEvent.DoesNotExist:
            raffle_event = None
            event = None
            org = None
            raffle_name = ''
        
        # Fallback to first winner's raffle event/org if query fails
        if not raffle_event and queryset.exists():
            first_winner = queryset.first()
            if first_winner and first_winner.prize and first_winner.prize.raffle_event:
                raffle_event = first_winner.prize.raffle_event
                event = raffle_event.event if hasattr(raffle_event, 'event') else None
                org = raffle_event.org if hasattr(raffle_event, 'org') else None
                raffle_name = raffle_event.name
        
        # Use raffle_name if it exists, otherwise use event_name (avoid duplication)
        display_name = raffle_name if raffle_name else (event.name if event else '')
        
        # Get prize name if filtering by prize
        prize_name = ''
        if prize_id:
            try:
                prize = Prize.objects.get(id=prize_id)
                prize_name = prize.name
            except Prize.DoesNotExist:
                pass
        
        # Prepare data for template with rank (ลำดับรางวัล)
        winners_data = []
        for rank, winner in enumerate(queryset, start=1):
            winners_data.append({
                'rank': rank,  # ลำดับรางวัล (1-based)
                'id': winner.id,  # Keep ID for reference
                'participant_name': winner.participant.name if winner.participant else '',
                'hospital_id': winner.participant.hospital_id if winner.participant and winner.participant.hospital_id else '',
                'prize_name': winner.prize.name if winner.prize else '',
                'department': winner.participant.department.name if winner.participant and winner.participant.department else '',
            })
        
        # Render HTML template
        html_content = render_to_string('raffle/winners_print.html', {
            'winners': winners_data,
            'raffle_name': display_name,  # Use display_name to avoid duplication
            'event_name': '',  # Don't show event_name separately
            'org_name': org.name if org else '',
            'search_query': search,
            'prize_filter': prize_name,
            'print_date': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
            'total_count': len(winners_data),
        })
        
        # Return HTML response (frontend will handle printing)
        response = HttpResponse(html_content, content_type='text/html; charset=utf-8')
        return response
    
    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_winners_excel(self, request):
        """Export winners to Excel"""
        org_id = getattr(request, 'org_id', None)
        if not org_id:
            return Response(
                {'error': 'Organization ID not found'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        raffle_event_id = request.query_params.get('raffle_event')
        if not raffle_event_id:
            return Response(
                {'error': 'Raffle Event ID is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get filter parameters
        search = request.query_params.get('search', '')
        prize_id = request.query_params.get('prize')
        only_unprinted = request.query_params.get('only_unprinted', 'false').lower() == 'true'
        start_index = request.query_params.get('start_index')
        end_index = request.query_params.get('end_index')
        
        # Get queryset (same logic as export_pdf)
        queryset = RaffleParticipant.objects.select_related(
            'prize', 'prize__raffle_event', 'participant', 'participant__department'
        ).filter(prize__raffle_event_id=raffle_event_id, prize__raffle_event__org_id=org_id)
        
        if search:
            queryset = queryset.filter(participant__name__icontains=search)
        
        if prize_id:
            queryset = queryset.filter(prize_id=prize_id)
        
        # Filter by print status
        if only_unprinted:
            queryset = queryset.filter(is_printed=False)
        
        # Order by selected_at descending (newest first = rank 1)
        queryset = queryset.order_by('-selected_at')
        
        # Apply range selection if provided
        if start_index or end_index:
            start = int(start_index) - 1 if start_index else 0  # Convert to 0-based index
            end = int(end_index) if end_index else None  # end_index is inclusive, so use as-is for slicing
            queryset = queryset[start:end]
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "รายชื่อผู้ได้รับรางวัล"
        
        # Set header row
        headers = ['ลำดับรางวัล', 'ID โรงพยาบาล', 'ชื่อผู้ได้รับรางวัล', 'รางวัล', 'หน่วยงาน', 'ลงชื่อ']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data with rank (ลำดับรางวัล)
        for rank, winner in enumerate(queryset, start=1):
            row_idx = rank + 1  # +1 for header row
            ws.cell(row=row_idx, column=1, value=rank)  # ลำดับรางวัล
            ws.cell(row=row_idx, column=2, value=winner.participant.hospital_id if winner.participant and winner.participant.hospital_id else '')  # ID โรงพยาบาล
            ws.cell(row=row_idx, column=3, value=winner.participant.name if winner.participant else '')  # ชื่อผู้ได้รับรางวัล
            ws.cell(row=row_idx, column=4, value=winner.prize.name if winner.prize else '')  # รางวัล
            ws.cell(row=row_idx, column=5, value=winner.participant.department.name if winner.participant and winner.participant.department else '')  # หน่วยงาน
            ws.cell(row=row_idx, column=6, value='')  # ลงชื่อ - ว่างไว้ให้เซ็น
        
        # Auto-adjust column widths
        column_widths = [12, 15, 35, 25, 20, 15]  # ลำดับรางวัล, ID โรงพยาบาล, Name, Prize, Department, ลงชื่อ
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Generate filename
        filename = f'winners_{raffle_event_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook to response
        wb.save(response)
        
        return response
    
    @action(detail=False, methods=['get'], url_path='public-list', permission_classes=[])
    def public_list(self, request):
        """List winners with pagination, filtering, and search (public, no auth required)"""
        from config.pagination import StandardResultsSetPagination
        
        # Get filter parameters
        raffle_event_id = request.query_params.get('raffle_event')
        prize_id = request.query_params.get('prize')
        search = request.query_params.get('search', '').strip()
        
        # Build queryset
        queryset = RaffleParticipant.objects.select_related(
            'prize', 'prize__raffle_event', 'participant', 'participant__department'
        )
        
        # Filter by raffle_event
        if raffle_event_id:
            try:
                queryset = queryset.filter(prize__raffle_event_id=raffle_event_id)
            except ValueError:
                return Response(
                    {'error': 'Invalid raffle_event parameter'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Filter by prize
        if prize_id:
            try:
                queryset = queryset.filter(prize_id=prize_id)
            except ValueError:
                return Response(
                    {'error': 'Invalid prize parameter'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Search by participant name
        if search:
            queryset = queryset.filter(participant__name__icontains=search)
        
        # Order by selected_at descending
        queryset = queryset.order_by('-selected_at')
        
        # Apply pagination
        paginator = StandardResultsSetPagination()
        page = paginator.paginate_queryset(queryset, request)
        
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return paginator.get_paginated_response(serializer.data)
        
        # If no pagination, return all results
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'count': queryset.count(),
            'results': serializer.data
        })


def get_or_create_participant(org_id, event_id, name, department_id=None, team_id=None):
    """
    Helper function to get or create a Participant
    Returns: (participant, created)
    """
    defaults = {}
    if department_id:
        defaults['department_id'] = department_id
    
    participant, created = Participant.objects.get_or_create(
        org_id=org_id,
        event_id=event_id,
        name=name.strip(),
        defaults=defaults
    )
    
    # Update department if provided and participant already exists
    if not created and department_id and participant.department_id != department_id:
        participant.department_id = department_id
        participant.save(update_fields=['department_id'])
    
    # Update team in metadata if provided
    if team_id:
        metadata = participant.metadata or {}
        metadata['team_id'] = team_id
        participant.metadata = metadata
        participant.save(update_fields=['metadata'])
    
    return participant, created


class RaffleControlViewSet(viewsets.ViewSet):
    """ViewSet for controlling raffle display via WebSocket"""
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['post'], url_path='spin')
    def control_spin(self, request):
        """Send spin command to display-test page via WebSocket"""
        raffle_event_id = request.data.get('raffle_event_id')
        prize_id = request.data.get('prize_id')
        display_count = request.data.get('display_count', 1)
        
        if not raffle_event_id or not prize_id:
            return Response(
                {'error': 'raffle_event_id and prize_id are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Broadcast control action via WebSocket
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        from datetime import datetime
        
        channel_layer = get_channel_layer()
        if channel_layer:
            # Send spin command
            async_to_sync(channel_layer.group_send)(
                f'raffle_{raffle_event_id}',
                {
                    'type': 'control_action',
                    'action': 'spin',
                    'data': {
                        'prize_id': prize_id,
                        'display_count': display_count
                    },
                    'timestamp': datetime.now().isoformat()
                }
            )
            # Send play sound command to display device
            async_to_sync(channel_layer.group_send)(
                f'raffle_{raffle_event_id}',
                {
                    'type': 'control_action',
                    'action': 'play_sound',
                    'data': {
                        'sound_file': 'wheel.mp3'
                    },
                    'timestamp': datetime.now().isoformat()
                }
            )
            # Send spin state to disable other devices
            async_to_sync(channel_layer.group_send)(
                f'raffle_{raffle_event_id}',
                {
                    'type': 'control_action',
                    'action': 'spin_state',
                    'data': {
                        'isSpinning': True
                    },
                    'timestamp': datetime.now().isoformat()
                }
            )
        
        return Response({'success': True, 'message': 'Spin command sent'})
    
    @action(detail=False, methods=['post'], url_path='save')
    def control_save(self, request):
        """Send save command to display-test page via WebSocket"""
        raffle_event_id = request.data.get('raffle_event_id')
        
        if not raffle_event_id:
            return Response(
                {'error': 'raffle_event_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Broadcast control action via WebSocket
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        from datetime import datetime
        
        channel_layer = get_channel_layer()
        if channel_layer:
            # Send save command
            async_to_sync(channel_layer.group_send)(
                f'raffle_{raffle_event_id}',
                {
                    'type': 'control_action',
                    'action': 'save',
                    'data': {},
                    'timestamp': datetime.now().isoformat()
                }
            )
            # Send spin state to re-enable other devices
            async_to_sync(channel_layer.group_send)(
                f'raffle_{raffle_event_id}',
                {
                    'type': 'control_action',
                    'action': 'spin_state',
                    'data': {
                        'isSpinning': False
                    },
                    'timestamp': datetime.now().isoformat()
                }
            )
        
        return Response({'success': True, 'message': 'Save command sent'})
    
    @action(detail=False, methods=['post'], url_path='select-prize')
    def control_select_prize(self, request):
        """Send select prize command to display-test page via WebSocket"""
        raffle_event_id = request.data.get('raffle_event_id')
        prize_id = request.data.get('prize_id')
        
        if not raffle_event_id or not prize_id:
            return Response(
                {'error': 'raffle_event_id and prize_id are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Broadcast control action via WebSocket
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        from datetime import datetime
        
        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f'raffle_{raffle_event_id}',
                {
                    'type': 'control_action',
                    'action': 'select_prize',
                    'data': {
                        'prize_id': prize_id
                    },
                    'timestamp': datetime.now().isoformat()
                }
            )
        
        return Response({'success': True, 'message': 'Select prize command sent'})
    
    @action(detail=False, methods=['post'], url_path='set-display-count')
    def control_set_display_count(self, request):
        """Send set display count command to display-test page via WebSocket"""
        raffle_event_id = request.data.get('raffle_event_id')
        display_count = request.data.get('display_count', 1)
        
        if not raffle_event_id:
            return Response(
                {'error': 'raffle_event_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Broadcast control action via WebSocket
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        from datetime import datetime
        
        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f'raffle_{raffle_event_id}',
                {
                    'type': 'control_action',
                    'action': 'set_display_count',
                    'data': {
                        'display_count': display_count
                    },
                    'timestamp': datetime.now().isoformat()
                }
            )
        
        return Response({'success': True, 'message': 'Set display count command sent'})
    
    @action(detail=False, methods=['get'], url_path='qr-code')
    def generate_qr_code(self, request):
        """Generate QR code URL for mobile controller"""
        raffle_event_id = request.query_params.get('raffle_event_id')
        
        if not raffle_event_id:
            return Response(
                {'error': 'raffle_event_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Generate mobile controller URL
        base_url = request.build_absolute_uri('/').rstrip('/')
        mobile_url = f"{base_url}/raffle/mobile-controller?raffle_event={raffle_event_id}"
        
        return Response({
            'success': True,
            'url': mobile_url,
            'qr_data': mobile_url
        })


class RaffleLogViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = RaffleLogSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['raffle_event', 'prize']
    ordering_fields = ['timestamp']
    ordering = ['-timestamp']
    permission_classes = [IsAuthenticated, IsStaffOrReadOnly]
    
    def get_queryset(self):
        org_id = getattr(self.request, 'org_id', None)
        if org_id:
            return RaffleLog.objects.filter(raffle_event__org_id=org_id)
        if self.request.user.is_superadmin():
            return RaffleLog.objects.all()
        return RaffleLog.objects.none()
    
    @action(detail=False, methods=['get'], url_path='export')
    def export_report(self, request):
        """Export raffle report"""
        raffle_event_id = request.query_params.get('raffle_event_id')
        if not raffle_event_id:
            return Response(
                {'error': 'raffle_event_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        org_id = getattr(request, 'org_id', None)
        if not org_id:
            return Response(
                {'error': 'Organization context required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            raffle_event = RaffleEvent.objects.get(id=raffle_event_id, org_id=org_id)
        except RaffleEvent.DoesNotExist:
            return Response(
                {'error': 'Raffle event not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        export_data = []
        for prize in raffle_event.prizes.all():
            for participant in prize.selected_participants.all():
                export_data.append({
                    'raffle_event': raffle_event.name,
                    'round': prize.round_number,
                    'prize': prize.name,
                    'participant': participant.participant.name,
                    'department': participant.participant.department.name if participant.participant.department else '',
                    'selected_at': participant.selected_at.isoformat(),
                    'seed': participant.seed_value
                })
        
        return Response({
            'success': True,
            'data': export_data,
            'count': len(export_data)
        })

