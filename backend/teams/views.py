from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
from django.template.loader import render_to_string
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from .models import Participant, Team, TeamMember
from .serializers import (
    ParticipantSerializer, TeamSerializer, TeamDetailSerializer, TeamMemberSerializer
)
from .algorithms import RandomAssignment, BalancedByDepartmentAssignment, RuleBasedAssignment
from core.permissions import IsOrgAdminOrReadOnly, IsStaffOrReadOnly
from core.utils import ImportProcessor, create_audit_log
from core.models import Organization
from rest_framework.permissions import IsAuthenticated
from config.pagination import StandardResultsSetPagination


class ParticipantViewSet(viewsets.ModelViewSet):
    serializer_class = ParticipantSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['org', 'event', 'department']
    search_fields = ['name']
    ordering_fields = ['hospital_id', 'name', 'created_at']
    ordering = ['hospital_id', 'name']
    permission_classes = [IsAuthenticated, IsOrgAdminOrReadOnly]
    pagination_class = StandardResultsSetPagination
    
    def get_queryset(self):
        org_id = getattr(self.request, 'org_id', None)
        if org_id:
            queryset = Participant.objects.filter(org_id=org_id)
            event_id = self.request.query_params.get('event')
            if event_id:
                queryset = queryset.filter(event_id=event_id)
            
            # Support filtering by department_name (custom filter)
            department_name = self.request.query_params.get('department_name')
            if department_name:
                queryset = queryset.filter(department__name=department_name)
            
            return queryset
        if self.request.user.is_superadmin():
            return Participant.objects.all()
        return Participant.objects.none()
    
    @action(detail=True, methods=['patch'], url_path='toggle-raffle-eligible')
    def toggle_raffle_eligible(self, request, pk=None):
        """Toggle raffle eligibility for a participant"""
        participant = self.get_object()
        is_eligible = request.data.get('is_raffle_eligible', not participant.is_raffle_eligible)
        
        participant.is_raffle_eligible = is_eligible
        participant.save(update_fields=['is_raffle_eligible'])
        
        serializer = self.get_serializer(participant)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], url_path='departments')
    def list_departments(self, request):
        """Get unique departments for participants in a specific event"""
        event_id = request.query_params.get('event')
        org_id = getattr(request, 'org_id', None)
        
        if not org_id:
            return Response(
                {'error': 'Organization context required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get queryset
        queryset = Participant.objects.filter(org_id=org_id)
        if event_id:
            queryset = queryset.filter(event_id=event_id)
        
        # Get unique departments from participants
        # Use .values() with .distinct() and then convert to list to ensure uniqueness
        departments = queryset.filter(
            department__isnull=False
        ).values('department__id', 'department__name').distinct()
        
        # Convert to set to ensure uniqueness (in case distinct() doesn't work as expected)
        # Then convert back to list
        unique_departments = {}
        for dept in departments:
            dept_id = dept['department__id']
            dept_name = dept['department__name']
            if dept_id and dept_name and dept_id not in unique_departments:
                unique_departments[dept_id] = dept_name
        
        # Format response
        department_list = [
            {'id': dept_id, 'name': dept_name}
            for dept_id, dept_name in unique_departments.items()
        ]
        
        # Sort by name
        department_list.sort(key=lambda x: x['name'])
        
        return Response(department_list)
    
    @action(detail=False, methods=['post'], url_path='reset-all-eligibility')
    def reset_all_eligibility(self, request):
        """Reset all participants' raffle eligibility to True for a specific event"""
        event_id = request.data.get('event_id')
        
        if not event_id:
            return Response(
                {'error': 'event_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        org_id = getattr(request, 'org_id', None)
        if not org_id:
            return Response(
                {'error': 'Organization context required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Update all participants in the event to have eligibility = True
        updated_count = Participant.objects.filter(
            org_id=org_id,
            event_id=event_id
        ).update(is_raffle_eligible=True)
        
        # Audit log
        from core.models import Organization
        org = Organization.objects.get(id=org_id)
        create_audit_log(
            user=request.user,
            org=org,
            action='bulk_update',
            model='Participant',
            changes={
                'event_id': event_id,
                'action': 'reset_all_eligibility',
                'updated_count': updated_count,
                'is_raffle_eligible': True
            },
            request=request
        )
        
        return Response({
            'success': True,
            'updated_count': updated_count,
            'message': f'Reset eligibility for {updated_count} participants'
        })
    
    def destroy(self, request, *args, **kwargs):
        """Override to handle delete with reason and audit log"""
        participant = self.get_object()
        org = participant.org
        
        # Get delete reason from request body
        reason = request.data.get('reason', '')
        if not reason or len(reason.strip()) < 10:
            return Response(
                {'detail': 'กรุณาระบุเหตุผลในการลบ (อย่างน้อย 10 ตัวอักษร)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Store participant data before deletion
        participant_data = {
            'id': participant.id,
            'name': participant.name,
            'event': participant.event.name if participant.event else None,
            'org': participant.org.name if participant.org else None,
            'department': participant.department.name if participant.department else None,
        }
        
        # Create audit log before deletion
        from core.utils import create_audit_log
        try:
            create_audit_log(
                user=request.user,
                org=org,
                action='delete',
                model='Participant',
                object_id=participant.id,
                changes={
                    'deleted_item': participant_data,
                    'delete_reason': reason.strip(),
                },
                request=request
            )
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Failed to create audit log for Participant {participant.id}: {e}')
        
        # Delete the participant
        self.perform_destroy(participant)
        
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=False, methods=['post'], url_path='import')
    def import_participants(self, request):
        """Import participants from file"""
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        event_id = request.data.get('event_id')
        if not event_id:
            return Response(
                {'error': 'event_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        org_id = getattr(request, 'org_id', None)
        if not org_id:
            return Response(
                {'error': 'Organization context required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        
        # Default column mapping for hospital Excel/CSV format
        default_mapping = {
            'hospital_id': 'ID',
            'first_name': 'ชื่อ',
            'last_name': 'นามสกุล',
            'department': 'หน่วยงาน'
        }
        
        # Get and parse column_mapping
        import json
        column_mapping_raw = request.data.get('column_mapping')
        
        if column_mapping_raw is None:
            column_mapping = default_mapping
        elif isinstance(column_mapping_raw, str):
            try:
                column_mapping = json.loads(column_mapping_raw)
                # Ensure it's a dict
                if not isinstance(column_mapping, dict):
                    column_mapping = default_mapping
            except (json.JSONDecodeError, TypeError):
                column_mapping = default_mapping
        elif isinstance(column_mapping_raw, dict):
            column_mapping = column_mapping_raw
        else:
            column_mapping = default_mapping
        
        # Ensure column_mapping is a dict
        if not isinstance(column_mapping, dict):
            column_mapping = default_mapping
        
        processor = ImportProcessor(file)
        read_result = processor.read_file()
        
        if not read_result.get('success', False):
            return Response(read_result, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate mapping
        validation = processor.validate_mapping(column_mapping)
        if not validation['success']:
            error_response = {
                'success': False,
                'error': 'Column mapping validation failed',
                'errors': validation.get('errors', []),
                'available_columns': read_result.get('columns', []),
                'expected_columns': list(column_mapping.values()),
            }
            return Response(error_response, status=status.HTTP_400_BAD_REQUEST)
        
        # Process data
        def validate_row(row_data):
            errors = []
            # At least first_name or last_name must be provided
            first_name = (row_data.get('first_name') or '').strip()
            last_name = (row_data.get('last_name') or '').strip()
            if not first_name and not last_name:
                errors.append('ต้องระบุชื่อหรือนามสกุลอย่างน้อย 1 อย่าง')
            # If hospital_id is provided, it must not be empty
            if 'hospital_id' in row_data and row_data.get('hospital_id') == '':
                errors.append('hospital_id cannot be empty if provided')
            return len(errors) == 0, errors
        
        result = processor.process_data(column_mapping, validate_row)
        
        if not result['success']:
            error_response = {
                'success': False,
                'error': 'Data validation failed',
                'errors': result.get('errors', []),
                'total_rows': result.get('total_rows', 0),
                'valid_rows': result.get('valid_rows', 0),
                'error_rows': result.get('error_rows', 0),
            }
            return Response(error_response, status=status.HTTP_400_BAD_REQUEST)
        
        # Create/Update participants
        created_count = 0
        updated_count = 0
        from core.models import Event, Department
        
        try:
            event = Event.objects.get(id=event_id, org_id=org_id)
        except Event.DoesNotExist:
            return Response(
                {'error': 'Event not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        for row_data in result['data']:
            # Combine first_name and last_name
            first_name = (row_data.get('first_name') or '').strip()
            last_name = (row_data.get('last_name') or '').strip()
            full_name = f"{first_name} {last_name}".strip()
            
            if not full_name:
                continue  # Skip if no name
            
            # Auto-create or find department
            department = None
            if row_data.get('department'):
                department_name = str(row_data['department']).strip()
                if department_name:
                    department, _ = Department.objects.get_or_create(
                        org_id=org_id,
                        name=department_name,
                        defaults={'is_active': True}
                    )
            
            # Get hospital_id if provided and convert to integer
            hospital_id = None
            hospital_id_raw = row_data.get('hospital_id')
            if hospital_id_raw:
                hospital_id_str = str(hospital_id_raw).strip()
                if hospital_id_str:
                    try:
                        hospital_id = int(hospital_id_str)
                    except (ValueError, TypeError):
                        # Skip invalid hospital_id, but don't fail the import
                        hospital_id = None
            
            # Store other fields in metadata
            metadata = {k: v for k, v in row_data.items() if k not in ['first_name', 'last_name', 'department', 'hospital_id']}
            
            # Update or create participant
            participant, created = Participant.objects.update_or_create(
                org_id=org_id,
                event=event,
                name=full_name,
                defaults={
                    'hospital_id': hospital_id,
                    'department': department,
                    'metadata': metadata
                }
            )
            if created:
                created_count += 1
            else:
                updated_count += 1
        
        # Audit log
        from core.models import Organization
        org = Organization.objects.get(id=org_id)
        create_audit_log(
            user=request.user,
            org=org,
            action='import',
            model='Participant',
            changes={
                'event_id': event_id,
                'created_count': created_count,
                'updated_count': updated_count,
                'total_rows': result['total_rows']
            },
            request=request
        )
        
        return Response({
            'success': True,
            'created_count': created_count,
            'updated_count': updated_count,
            'total_rows': result['total_rows']
        })
    
    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_participants_excel(self, request):
        """Export participants to Excel using template"""
        org_id = getattr(request, 'org_id', None)
        if not org_id:
            return Response(
                {'error': 'Organization ID not found'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        event_id = request.query_params.get('event')
        if not event_id:
            return Response(
                {'error': 'Event ID is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get filter parameters
        search = request.query_params.get('search', '')
        department_name = request.query_params.get('department_name', '')
        
        # Get queryset (same logic as list, but without pagination)
        queryset = Participant.objects.filter(org_id=org_id, event_id=event_id)
        
        if search:
            queryset = queryset.filter(name__icontains=search)
        
        if department_name:
            queryset = queryset.filter(department__name=department_name)
        
        # Order by hospital_id, then name
        queryset = queryset.order_by('hospital_id', 'name').select_related('department')
        
        # Load template
        # Template path: backend/excel_templates/template1.xlsx
        BASE_DIR = Path(__file__).resolve().parent.parent
        template_path = BASE_DIR / 'excel_templates' / 'template1.xlsx'
        
        if not template_path.exists():
            return Response(
                {'error': 'Template file not found'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        # Load workbook from template
        wb = load_workbook(template_path)
        ws = wb.active
        
        # Check existing headers in row 1
        existing_headers = []
        for col in range(1, 10):  # Check up to 10 columns
            header_value = ws.cell(row=1, column=col).value
            if header_value:
                existing_headers.append(str(header_value))
            else:
                break
        
        # Determine signature column
        signature_col = 4  # Default to column D
        
        # If no headers exist, create them
        if not existing_headers:
            headers = ['ID โรงพยาบาล', 'ชื่อ-นามสกุล', 'หน่วยงาน', 'ลงชื่อ']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            signature_col = 4
        # If headers exist but "ลงชื่อ" column is missing, add it
        elif 'ลงชื่อ' not in existing_headers:
            # Find the last column with header
            last_col = len(existing_headers) + 1
            cell = ws.cell(row=1, column=last_col, value='ลงชื่อ')
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            signature_col = last_col
        else:
            # Find the column index of "ลงชื่อ"
            for idx, header in enumerate(existing_headers, start=1):
                if 'ลงชื่อ' in str(header):
                    signature_col = idx
                    break
        
        # Data starts from row 2 (row 1 is header)
        start_row = 2
        
        # Write data
        # Column A = hospital_id
        # Column B = name
        # Column C = department
        # signature_col = ลงชื่อ (empty for signature)
        for idx, participant in enumerate(queryset, start=start_row):
            ws.cell(row=idx, column=1, value=participant.hospital_id if participant.hospital_id else '')
            ws.cell(row=idx, column=2, value=participant.name)
            ws.cell(row=idx, column=3, value=participant.department.name if participant.department else '')
            ws.cell(row=idx, column=signature_col, value='')  # ลงชื่อ - ว่างไว้ให้เซ็น
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Generate filename
        filename = f'participants_{event_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook to response
        wb.save(response)
        
        return response
    
    @action(detail=False, methods=['get'], url_path='export-pdf')
    def export_participants_pdf(self, request):
        """Export participants to PDF for printing"""
        org_id = getattr(request, 'org_id', None)
        if not org_id:
            return Response(
                {'error': 'Organization ID not found'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        event_id = request.query_params.get('event')
        if not event_id:
            return Response(
                {'error': 'Event ID is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get filter parameters
        search = request.query_params.get('search', '')
        department_name = request.query_params.get('department_name', '')
        
        # Get queryset (same logic as list, but without pagination)
        queryset = Participant.objects.filter(org_id=org_id, event_id=event_id)
        
        if search:
            queryset = queryset.filter(name__icontains=search)
        
        if department_name:
            queryset = queryset.filter(department__name=department_name)
        
        # Order by hospital_id, then name
        queryset = queryset.order_by('hospital_id', 'name').select_related('department', 'event', 'org')
        
        # Get event and org info
        from core.models import Event
        try:
            event = Event.objects.get(id=event_id, org_id=org_id)
            org = event.org if hasattr(event, 'org') else None
        except Event.DoesNotExist:
            event = None
            org = None
        
        # Fallback to first participant's event/org if event query fails
        if not event and queryset.exists():
            first_participant = queryset.first()
            event = first_participant.event
            org = first_participant.org
        
        # Prepare data for template
        participants_data = []
        for participant in queryset:
            participants_data.append({
                'hospital_id': participant.hospital_id if participant.hospital_id else '',
                'name': participant.name,
                'department': participant.department.name if participant.department else '',
            })
        
        # Render HTML template
        html_content = render_to_string('teams/participants_print.html', {
            'participants': participants_data,
            'event_name': event.name if event else '',
            'org_name': org.name if org else '',
            'search_query': search,
            'department_filter': department_name,
            'print_date': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
            'total_count': len(participants_data),
        })
        
        # Return HTML response (frontend will handle printing)
        response = HttpResponse(html_content, content_type='text/html; charset=utf-8')
        return response


class TeamViewSet(viewsets.ModelViewSet):
    serializer_class = TeamSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['org', 'event']
    search_fields = ['color_name']
    ordering_fields = ['color_name', 'created_at']
    ordering = ['color_name']
    permission_classes = [IsAuthenticated, IsOrgAdminOrReadOnly]
    
    def get_queryset(self):
        org_id = getattr(self.request, 'org_id', None)
        if org_id:
            queryset = Team.objects.filter(org_id=org_id)
            event_id = self.request.query_params.get('event')
            if event_id:
                queryset = queryset.filter(event_id=event_id)
            return queryset
        if self.request.user.is_superadmin():
            return Team.objects.all()
        return Team.objects.none()
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return TeamDetailSerializer
        return TeamSerializer
    
    def perform_create(self, serializer):
        """Override to add audit log on create"""
        team = serializer.save()
        org_id = getattr(self.request, 'org_id', None)
        org = team.org if team.org else (Organization.objects.get(id=org_id) if org_id else None)
        
        create_audit_log(
            user=self.request.user,
            org=org,
            action='create',
            model='Team',
            object_id=team.id,
            changes={'color_name': team.color_name, 'color_code': team.color_code},
            request=self.request
        )
    
    def perform_update(self, serializer):
        """Override to add audit log on update"""
        old_team = self.get_object()
        team = serializer.save()
        org = team.org
        
        # Track changes
        changes = {}
        if old_team.color_name != team.color_name:
            changes['color_name'] = {'old': old_team.color_name, 'new': team.color_name}
        if old_team.color_code != team.color_code:
            changes['color_code'] = {'old': old_team.color_code, 'new': team.color_code}
        if old_team.max_members != team.max_members:
            changes['max_members'] = {'old': old_team.max_members, 'new': team.max_members}
        
        create_audit_log(
            user=self.request.user,
            org=org,
            action='update',
            model='Team',
            object_id=team.id,
            changes=changes,
            request=self.request
        )
    
    def destroy(self, request, *args, **kwargs):
        """Override to require reason and add audit log"""
        reason = request.data.get('reason', '')
        if not reason or len(reason.strip()) < 10:
            return Response(
                {'error': 'Reason is required and must be at least 10 characters'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        team = self.get_object()
        org = team.org
        team_data = {
            'color_name': team.color_name,
            'color_code': team.color_code,
            'event': team.event.name if team.event else None,
        }
        
        # Create audit log before deletion
        create_audit_log(
            user=request.user,
            org=org,
            action='delete',
            model='Team',
            object_id=team.id,
            changes={'reason': reason, 'deleted_data': team_data},
            request=request
        )
        
        # Delete the team
        self.perform_destroy(team)
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=False, methods=['post'], url_path='assign')
    def assign_teams(self, request):
        """Assign participants to teams using algorithm"""
        event_id = request.data.get('event_id')
        algorithm = request.data.get('algorithm', 'random')  # random, balanced, rule-based
        rules = request.data.get('rules', {})
        
        if not event_id:
            return Response(
                {'error': 'event_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        org_id = getattr(request, 'org_id', None)
        if not org_id:
            return Response(
                {'error': 'Organization context required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get participants and teams
        participants = Participant.objects.filter(org_id=org_id, event_id=event_id)
        teams = Team.objects.filter(org_id=org_id, event_id=event_id)
        
        if not teams.exists():
            return Response(
                {'error': 'No teams found for this event'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Choose algorithm
        if algorithm == 'random':
            alg = RandomAssignment(list(participants), list(teams))
        elif algorithm == 'balanced':
            alg = BalancedByDepartmentAssignment(list(participants), list(teams))
        elif algorithm == 'rule-based':
            alg = RuleBasedAssignment(list(participants), list(teams), rules)
        else:
            return Response(
                {'error': f'Unknown algorithm: {algorithm}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Assign
        assignments = alg.assign()
        
        # Create TeamMember records
        created_count = 0
        for assignment in assignments:
            team_member, created = TeamMember.objects.get_or_create(
                team=assignment['team'],
                participant=assignment['participant'],
                defaults={'is_moved': False}
            )
            if created:
                created_count += 1
        
        # Audit log
        from core.models import Organization
        org = Organization.objects.get(id=org_id)
        create_audit_log(
            user=request.user,
            org=org,
            action='create',
            model='TeamMember',
            changes={'event_id': event_id, 'algorithm': algorithm, 'created_count': created_count},
            request=request
        )
        
        return Response({
            'success': True,
            'assigned_count': created_count,
            'total_participants': len(assignments)
        })


class TeamMemberViewSet(viewsets.ModelViewSet):
    serializer_class = TeamMemberSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['team', 'participant', 'is_pinned', 'is_moved']
    search_fields = ['participant__name']
    permission_classes = [IsAuthenticated, IsStaffOrReadOnly]
    
    def get_queryset(self):
        org_id = getattr(self.request, 'org_id', None)
        if org_id:
            return TeamMember.objects.filter(team__org_id=org_id)
        if self.request.user.is_superadmin():
            return TeamMember.objects.all()
        return TeamMember.objects.none()
    
    def perform_create(self, serializer):
        """Override to auto-set event from team and add audit log"""
        team_member = serializer.save()
        # Auto-set event from team if not provided (TeamMember.save() should handle this, but ensure it here too)
        if not team_member.event_id and team_member.team_id:
            team_member.event = team_member.team.event
            team_member.save(update_fields=['event'])
        
        # Audit log
        org = team_member.team.org if team_member.team else None
        create_audit_log(
            user=self.request.user,
            org=org,
            action='create',
            model='TeamMember',
            object_id=team_member.id,
            changes={
                'team_id': team_member.team_id,
                'participant_id': team_member.participant_id,
                'event_id': team_member.event_id
            },
            request=self.request
        )
    
    @action(detail=True, methods=['post'], url_path='move')
    def move_member(self, request, pk=None):
        """Move member to another team"""
        team_member = self.get_object()
        new_team_id = request.data.get('team_id')
        
        if not new_team_id:
            return Response(
                {'error': 'team_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            new_team = Team.objects.get(id=new_team_id, org_id=team_member.team.org_id)
        except Team.DoesNotExist:
            return Response(
                {'error': 'Team not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        old_team = team_member.team
        team_member.team = new_team
        team_member.is_moved = True
        team_member.moved_from_team = old_team
        team_member.save()
        
        # Audit log
        create_audit_log(
            user=request.user,
            org=team_member.team.org,
            action='update',
            model='TeamMember',
            object_id=team_member.id,
            changes={'moved_from': old_team.id, 'moved_to': new_team.id},
            request=request
        )
        
        serializer = self.get_serializer(team_member)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'], url_path='pin')
    def pin_member(self, request, pk=None):
        """Pin/unpin member (lock/unlock)"""
        team_member = self.get_object()
        is_pinned = request.data.get('is_pinned', True)
        
        team_member.is_pinned = is_pinned
        team_member.save()
        
        # Audit log
        create_audit_log(
            user=request.user,
            org=team_member.team.org,
            action='update',
            model='TeamMember',
            object_id=team_member.id,
            changes={'is_pinned': is_pinned},
            request=request
        )
        
        serializer = self.get_serializer(team_member)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], url_path='export')
    def export_teams(self, request):
        """Export team members to CSV/Excel"""
        event_id = request.query_params.get('event_id')
        if not event_id:
            return Response(
                {'error': 'event_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        org_id = getattr(request, 'org_id', None)
        if not org_id:
            return Response(
                {'error': 'Organization context required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        teams = Team.objects.filter(org_id=org_id, event_id=event_id)
        export_data = []
        
        for team in teams:
            for member in team.members.all():
                export_data.append({
                    'team': team.color_name,
                    'team_color': team.color_code,
                    'participant': member.participant.name,
                    'department': member.participant.department.name if member.participant.department else '',
                    'is_pinned': member.is_pinned,
                    'is_moved': member.is_moved
                })
        
        return Response({
            'success': True,
            'data': export_data,
            'count': len(export_data)
        })

